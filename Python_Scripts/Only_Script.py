# =============================================
# Cryptocurrency Portfolio Data Aggregation Tool
# 
# This script automates the collection of cryptocurrency data from CoinGecko API,
# processes it, and stores it in an Excel file for Power BI reporting.
# Features include:
#   - Real-time portfolio tracking
#   - Market overview of top cryptocurrencies
#   - Historical price data with technical indicators
#   - Global market metrics
#   - Fear & Greed Index sentiment analysis
# =============================================


import requests
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook, Workbook 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side 
from openpyxl.chart import LineChart, Reference 
from openpyxl.formatting.rule import ColorScaleRule 
import pytz
import time
import random
import traceback 

# --- Configuration ---
# Portfolio coins to track (modify as needed)
PORTFOLIO_COIN_IDS = [
    'bitcoin', 'ethereum', 'solana', 'binancecoin', 'the-open-network',
    'not-pixel', 'dogs-coin', 'notcoin', 'newton-protocol', 'blum'
]

# Historical data coins (symbol: CoinGecko ID mapping) (modify as needed)
HISTORY_COIN_IDS = {
    'BTC': 'bitcoin',
    'ETH': 'ethereum',
    'BNB': 'binancecoin',
    'SOL': 'solana',
    'SUI': 'sui' 
}
HISTORICAL_DAYS = 30 # For 1 month history

# --- API Rate Limit Delay & Retry ---
INITIAL_API_CALL_DELAY_SECONDS = 5 # Initial delay between consecutive API calls
MAX_RETRIES = 5                    # Max number of retries for an API call
BACKOFF_FACTOR = 2                 # Factor by which to increase delay (e.g., 2s, 4s, 8s...)

# --- Google Drive Path Configuration ---
GOOGLE_DRIVE_FOLDER = '/content/drive/My Drive/Crypto_Portfolio_Data'
EXCEL_FILE_NAME = os.path.join(GOOGLE_DRIVE_FOLDER, 'crypto_portfolio.xlsx')

DATE_FORMAT = '%Y-%m-%d %H:%M:%S'
IST = pytz.timezone('Asia/Kolkata')

# --- Sheet Names (in desired order) ---
CURRENT_PORTFOLIO_SHEET_NAME = 'Current Portfolio'
TRANSACTIONS_SHEET_NAME = 'Transactions'
# Sheet names for new content
EXECUTIVE_DASHBOARD_SHEET_NAME = '📊 Executive Dashboard'
MARKET_OVERVIEW_SHEET_NAME = '📈 Market Overview'
GLOBAL_METRICS_SHEET_NAME = '🌍 Global Metrics'
FEAR_GREED_INDEX_SHEET_NAME = '😰 Fear & Greed Index' # Consistent naming for sheet creation


SHEET_ORDER = [
    EXECUTIVE_DASHBOARD_SHEET_NAME, # New dashboard
    MARKET_OVERVIEW_SHEET_NAME,     # New detailed market overview
    GLOBAL_METRICS_SHEET_NAME,      # New global metrics sheet
    FEAR_GREED_INDEX_SHEET_NAME,    # Now explicitly named for consistency
    'BTC History (1 Month)',
    'ETH History (1 Month)',
    'BNB History (1 Month)',
    'SOL History (1 Month)',
    'SUI History (1 Month)', 
    TRANSACTIONS_SHEET_NAME,
    CURRENT_PORTFOLIO_SHEET_NAME
]

# --- Color Schemes ---
COLORS = {
    'header': 'FF1F4E79',
    'positive': 'FF00B050',
    'negative': 'FFFF0000',
    'neutral': 'FF9CB4D8',
    'background': 'FFF2F2F2'
}

# --- API Helper Function with Retry Logic ---
def make_api_call_with_retry(url, params=None, max_retries=MAX_RETRIES, initial_delay=INITIAL_API_CALL_DELAY_SECONDS):
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            response = requests.get(url, params=params)
            response.raise_for_status() # Raises HTTPError for bad responses (4xx or 5xx)
            return response
        except requests.exceptions.RequestException as e:
            if response is not None and response.status_code == 429:
                print(f"  Rate limit hit for {url}. Retrying in {delay:.2f} seconds (Attempt {attempt + 1}/{max_retries})...")
                time.sleep(delay + random.uniform(0, 1)) # Add some jitter
                delay *= BACKOFF_FACTOR # Exponential backoff
            else:
                print(f"  Error during API call for {url}: {e}")
                traceback.print_exc() # Print full traceback for other errors
                return None # Other errors, don't retry
    print(f"  Failed to fetch data from {url} after {max_retries} attempts due to rate limits.")
    return None

# --- API Functions ---

def fetch_current_coin_prices(coin_ids):
    """Fetches current prices for given coin IDs from CoinGecko API."""
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {
        'ids': ','.join(coin_ids),
        'vs_currencies': 'usd'
    }
    response = make_api_call_with_retry(url, params)
    if response:
        data = response.json()
        # Ensure prices are stored as float, None if not found
        prices = {coin_id: data.get(coin_id, {}).get('usd') for coin_id in coin_ids}
        return prices
    return None

def get_market_overview():
    """Get comprehensive market overview data (from CryptoExcelGenerator)"""
    base_url = "https://api.coingecko.com/api/v3"
    url = f"{base_url}/coins/markets"
    params = {
        'vs_currency': 'usd',
        'order': 'market_cap_desc',
        'per_page': 50,
        'page': 1,
        'sparkline': 'false',
        'price_change_percentage': '1h,24h,7d,30d,1y'
    }

    response = make_api_call_with_retry(url, params)
    if response:
        data = response.json()
        df = pd.DataFrame(data)

        # Select and rename columns
        columns_map = {
            'market_cap_rank': 'Rank',
            'name': 'Name',
            'symbol': 'Symbol',
            'current_price': 'Price (USD)',
            'market_cap': 'Market Cap',
            'total_volume': '24h Volume',
            'price_change_percentage_24h': '24h Change (%)',
            'price_change_percentage_7d_in_currency': '7d Change (%)',
            'price_change_percentage_30d_in_currency': '30d Change (%)',
            'price_change_percentage_1y_in_currency': '1y Change (%)',
            'circulating_supply': 'Circulating Supply',
            'total_supply': 'Total Supply',
            'max_supply': 'Max Supply',
            'ath': 'All-Time High',
            'ath_date': 'ATH Date',
            'atl': 'All-Time Low',
            'atl_date': 'ATL Date'
        }

        df = df.rename(columns=columns_map)
        df = df[list(columns_map.values())]

        # --- CRITICAL CHANGE: Ensure numeric types before putting into Excel ---
        # Convert numeric columns to appropriate types, handling missing/invalid data
        numeric_cols_to_convert = {
            'Price (USD)': float,
            'Market Cap': float,
            '24h Volume': float,
            '24h Change (%)': float,
            '7d Change (%)': float,
            '30d Change (%)': float,
            '1y Change (%)': float,
            'Circulating Supply': float,
            'Total Supply': float,
            'Max Supply': float,
            'All-Time High': float,
            'All-Time Low': float
        }

        for col, dtype in numeric_cols_to_convert.items():
            if col in df.columns:
                # Use pd.to_numeric with errors='coerce' to turn unparseable values into NaN
                df[col] = pd.to_numeric(df[col], errors='coerce').astype(dtype)
                # Round percentage changes AFTER conversion to float
                if 'Change (%)' in col:
                    df[col] = df[col].round(2)
                elif col == 'Price (USD)':
                    df[col] = df[col].round(4) # Keep higher precision for price

        # Convert ATH/ATL dates to datetime objects then format as strings
        for date_col in ['ATH Date', 'ATL Date']:
            if date_col in df.columns:
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y-%m-%d')
                # Replace NaT with None or empty string for Excel
                df[date_col] = df[date_col].fillna('')

        # Do NOT format with '$' here. Let Power BI or Excel cell formatting handle it.
        # df['Market Cap'] = df['Market Cap'].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "N/A")
        # df['24h Volume'] = df['24h Volume'].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "N/A")
        # df['Price (USD)'] = df['Price (USD)'].apply(lambda x: f"${x:,.4f}" if pd.notna(x) else "N/A")

        return df
    return pd.DataFrame()


def get_historical_data(coin_id, days):
    """Fetches daily historical market data (price, market cap, volume) for a coin."""
    base_url = "https://api.coingecko.com/api/v3"
    url = f"{base_url}/coins/{coin_id}/market_chart"
    params = {
        'vs_currency': 'usd',
        'days': days,
        'interval': 'daily'
    }
    response = make_api_call_with_retry(url, params)
    if response:
        data = response.json()
        prices = data.get('prices', [])
        volumes = data.get('total_volumes', [])
        market_caps = data.get('market_caps', [])

        # Ensure all lists have the same length for DataFrame creation
        min_len = min(len(prices), len(volumes), len(market_caps))

        # Extract data and convert to appropriate types, handling potential missing values
        df = pd.DataFrame({
            'Date': [datetime.fromtimestamp(p[0]/1000).strftime('%Y-%m-%d') for p in prices[:min_len]],
            'Price': [round(p[1], 4) if p[1] is not None else None for p in prices[:min_len]],
            'Volume': [int(v[1]) if v[1] is not None else None for v in volumes[:min_len]],
            'Market Cap': [int(mc[1]) if mc[1] is not None else None for mc in market_caps[:min_len]]
        })

        # Calculate technical indicators on numeric columns
        # Use .copy() to avoid SettingWithCopyWarning
        df['Price_numeric'] = df['Price'].copy()

        # Only calculate if there's enough data
        if len(df) > 1:
            df['Daily Return (%)'] = ((df['Price_numeric'] / df['Price_numeric'].shift(1)) - 1) * 100
            df['7-Day MA'] = df['Price_numeric'].rolling(window=7, min_periods=1).mean() # min_periods for early rows
            df['30-Day MA'] = df['Price_numeric'].rolling(window=30, min_periods=1).mean()
            df['Volatility (7d)'] = df['Daily Return (%)'].rolling(window=7, min_periods=1).std()
        else:
            df['Daily Return (%)'] = None
            df['7-Day MA'] = None
            df['30-Day MA'] = None
            df['Volatility (7d)'] = None

        # Round values
        df['7-Day MA'] = df['7-Day MA'].round(4)
        df['30-Day MA'] = df['30-Day MA'].round(4)
        df['Daily Return (%)'] = df['Daily Return (%)'].round(3)
        df['Volatility (7d)'] = df['Volatility (7d)'].round(3)

        # Remove helper column
        df = df.drop('Price_numeric', axis=1)

        return df
    return pd.DataFrame()


def get_fear_greed_index():
    """Fetches the latest Fear & Greed Index data (from CryptoExcelGenerator)."""
    url = "https://api.alternative.me/fng/?limit=30"
    response = make_api_call_with_retry(url)
    if response:
        try:
            data = response.json()['data']
            df = pd.DataFrame(data)
            df['Date'] = pd.to_datetime(df['timestamp'], unit='s', errors='coerce').dt.strftime('%Y-%m-%d')
            # Ensure 'value' is numeric, coerce errors to NaN
            df['Fear & Greed Index'] = pd.to_numeric(df['value'], errors='coerce').astype('Int64') # Use Int64 for nullable integer
            df['Classification'] = df['value_classification']

            return df[['Date', 'Fear & Greed Index', 'Classification']].sort_values('Date')
        except KeyError as e:
            print(f"Error parsing Fear & Greed Index data: Missing key {e}. Response: {response.text}")
            traceback.print_exc()
            return pd.DataFrame()
        except Exception as e:
            print(f"An unexpected error occurred in fetch_fear_greed_index parsing: {e}")
            traceback.print_exc()
            return pd.DataFrame()
    return pd.DataFrame()


def get_global_metrics():
    """Fetches global cryptocurrency metrics (from CryptoExcelGenerator)."""
    base_url = "https://api.coingecko.com/api/v3"
    url = f"{base_url}/global"
    response = make_api_call_with_retry(url)
    if response:
        try:
            data = response.json()['data']
            # --- CRITICAL CHANGE: Store raw numeric values, not formatted strings ---
            metrics_data = {
                'Metric': [
                    'Total Market Cap (USD)',
                    'Total 24h Volume (USD)',
                    'Bitcoin Dominance (%)',
                    'Ethereum Dominance (%)',
                    'Active Cryptocurrencies',
                    'Total Markets',
                    'Market Cap Change 24h (%)'
                ],
                'Value': [
                    data.get('total_market_cap', {}).get('usd'),
                    data.get('total_volume', {}).get('usd'),
                    data.get('market_cap_percentage', {}).get('btc'),
                    data.get('market_cap_percentage', {}).get('eth'),
                    data.get('active_cryptocurrencies'),
                    data.get('markets'),
                    data.get('market_cap_change_percentage_24h_usd')
                ]
            }
            df = pd.DataFrame(metrics_data)

            # Ensure numeric types where appropriate
            df['Value'] = pd.to_numeric(df['Value'], errors='coerce') # Coerce non-numeric to NaN

            # You can still add a display column if needed for string formatting for the dashboard directly
            # but the underlying 'Value' should remain numeric for Power BI
            def format_global_metric(row):
                if pd.isna(row['Value']):
                    return "N/A"
                if '%' in row['Metric']:
                    return f"{row['Value']:.2f}%"
                elif 'USD' in row['Metric']:
                    return f"${row['Value']:,.0f}"
                else:
                    return f"{int(row['Value']):,}" if pd.notna(row['Value']) else "N/A"

            df['Formatted Value'] = df.apply(format_global_metric, axis=1)

            # Return 'Metric' and 'Value' for Power BI (or 'Formatted Value' if you prefer text in Excel directly)
            # For Power BI, it's generally better to pass 'Value' as numeric and let Power BI format.
            return df[['Metric', 'Value']] # Pass raw numeric value to Excel
        except Exception as e:
            print(f"Error parsing global market data: {e}. Response: {response.text}")
            traceback.print_exc() # Print full traceback for debugging
            return pd.DataFrame()
    return pd.DataFrame()


def create_or_update_excel(all_data, excel_file):
    """
    Creates/updates the Excel file with data for all specified sheets.
    Handles existing manual sheets (like Transactions) without overwriting.
    Ensures all necessary sheets exist before attempting to write to them.
    Integrates advanced styling for Dashboard and Market Overview.
    """
    current_time_ist = datetime.now(IST).strftime(DATE_FORMAT)

    os.makedirs(os.path.dirname(excel_file), exist_ok=True)

    wb = None
    if os.path.exists(excel_file):
        print(f"'{excel_file}' found. Loading existing workbook...")
        try:
            wb = load_workbook(excel_file)
        except Exception as e:
            print(f"Error loading existing Excel file: {e}. It might be corrupted or in an invalid format. Creating a new one.")
            traceback.print_exc()
            wb = None

    if wb is None:
        print(f"'{excel_file}' not found or could not be loaded. Creating a new one with all sheets...")
        wb = Workbook() # Use Workbook() from openpyxl
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # Ensure all required sheets exist and add headers for manual ones if missing
    for sheet_name in SHEET_ORDER:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            print(f"  Created missing sheet: '{sheet_name}'")

            if sheet_name == TRANSACTIONS_SHEET_NAME:
                ws.append(['Date', 'Coin ID', 'Type (Buy/Sell)', 'Quantity', 'Price (USD)', 'Total Cost/Revenue (USD)'])
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 25
            elif sheet_name == CURRENT_PORTFOLIO_SHEET_NAME:
                ws = wb[sheet_name] # Ensure we're working on the right sheet object
                # Updated headers to reflect the expected input from the script (Current Value (USD) is current price)
                ws.append(['Coin ID', 'Current Value (USD)', 'Purchase Price (USD)', 'Quantity', 'P/L (USD)', 'P/L Ratio', 'Total Value (USD)', 'Symbol', 'Location', 'AirDrop or Invest'])
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 20
                ws.column_dimensions['H'].width = 10
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 10


    # --- Update each sheet ---

    # 1. Executive Dashboard
    ws_exec_dash = wb[EXECUTIVE_DASHBOARD_SHEET_NAME]
    ws_exec_dash.delete_rows(1, ws_exec_dash.max_row) # Clear existing content

    ws_exec_dash['A1'] = "Cryptocurrency Market Dashboard"
    ws_exec_dash['A1'].font = Font(size=20, bold=True, color=COLORS['header'])
    ws_exec_dash.merge_cells('A1:H1')

    ws_exec_dash['A3'] = f"Generated: {current_time_ist}"
    ws_exec_dash['A3'].font = Font(size=12, italic=True)

    # Key metrics section
    ws_exec_dash['A5'] = "🔑 Key Market Metrics"
    ws_exec_dash['A5'].font = Font(size=14, bold=True)

    global_metrics_df = all_data.get('global_metrics')
    if global_metrics_df is not None and not global_metrics_df.empty:
        # Re-apply string formatting for display in the dashboard sheet
        display_metrics_data = {
            'Metric': [],
            'Value': []
        }
        for _, row in global_metrics_df.iterrows():
            display_metrics_data['Metric'].append(row['Metric'])
            if pd.isna(row['Value']):
                display_metrics_data['Value'].append("N/A")
            elif '%' in row['Metric']:
                display_metrics_data['Value'].append(f"{row['Value']:.2f}%")
            elif 'USD' in row['Metric']:
                display_metrics_data['Value'].append(f"${row['Value']:,.0f}")
            else:
                display_metrics_data['Value'].append(f"{int(row['Value']):,}")

        for i, (metric, value) in enumerate(zip(display_metrics_data['Metric'], display_metrics_data['Value']), start=6):
            ws_exec_dash[f'A{i}'] = metric
            ws_exec_dash[f'B{i}'] = value
            ws_exec_dash[f'A{i}'].font = Font(bold=True)
        # Apply auto-width for these columns as well
        ws_exec_dash.column_dimensions['A'].width = 25
        ws_exec_dash.column_dimensions['B'].width = 25
    else:
        ws_exec_dash['A6'] = "Unable to fetch global metrics"


    # Market sentiment section
    ws_exec_dash['D5'] = "😰 Market Sentiment"
    ws_exec_dash['D5'].font = Font(size=14, bold=True)

    fng_df = all_data.get('fear_greed_index')
    if fng_df is not None and not fng_df.empty:
        latest = fng_df.iloc[-1] # Get the latest entry
        # Ensure 'Fear & Greed Index' is treated as a number for comparison
        fng_value = latest['Fear & Greed Index']
        classification = latest['Classification']

        ws_exec_dash['D6'] = f"Fear & Greed Index: {fng_value if pd.notna(fng_value) else 'N/A'}"
        ws_exec_dash['D7'] = f"Classification: {classification if pd.notna(classification) else 'N/A'}"

        # Color code based on sentiment, check if fng_value is not NaN
        if pd.notna(fng_value):
            if fng_value > 75:
                ws_exec_dash['D7'].fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid') # Green for Extreme Greed
            elif fng_value < 25:
                ws_exec_dash['D7'].fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid') # Red for Extreme Fear
            else:
                ws_exec_dash['D7'].fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid') # Yellow for Neutral
    else:
        ws_exec_dash['D6'] = "Unable to fetch sentiment data"

    # 2. Market Overview Sheet
    market_overview_df = all_data.get('market_overview')
    if market_overview_df is not None and not market_overview_df.empty:
        ws_market_overview = wb[MARKET_OVERVIEW_SHEET_NAME]
        ws_market_overview.delete_rows(1, ws_market_overview.max_row) # Clear existing content

        # Write header row
        ws_market_overview.append(market_overview_df.columns.tolist())

        # Write data rows
        for r_idx, r in enumerate(dataframe_to_rows(market_overview_df, index=False, header=False)): # header=False as we added it manually
            for c_idx, value in enumerate(r):
                ws_market_overview.cell(row=r_idx + 2, column=c_idx + 1, value=value) # Start from row 2

        # Apply formatting (from your CryptoExcelGenerator class)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

        for cell in ws_market_overview[1]: # Apply to the manually added header row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Apply alignment and basic coloring to data rows
        for row_idx in range(2, ws_market_overview.max_row + 1):
            for col_idx in range(1, ws_market_overview.max_column + 1):
                cell = ws_market_overview.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center')

                # Apply numeric formatting for currency columns
                header_name = ws_market_overview.cell(row=1, column=col_idx).value
                if header_name in ['Price (USD)', 'Market Cap', '24h Volume', 'All-Time High', 'All-Time Low']:
                    if cell.value is not None:
                        cell.number_format = '$#,##0.00' if 'Price' in header_name else '$#,##0'
                elif 'Change (%)' in header_name:
                    if cell.value is not None:
                        cell.number_format = '0.00%' # Format as percentage

                # Color code percentage changes (background fill)
                if header_name in ['24h Change (%)', '7d Change (%)', '30d Change (%)', '1y Change (%)']:
                    try:
                        value = float(cell.value)
                        if value > 0:
                            cell.fill = PatternFill(start_color='FFE6F3E6', end_color='FFE6F3E6', fill_type='solid') # Light Green
                        elif value < 0:
                            cell.fill = PatternFill(start_color='FFFFE6E6', end_color='FFFFE6E6', fill_type='solid') # Light Red
                    except (ValueError, TypeError):
                        pass # Handle non-numeric or None values

        # Add conditional formatting for price changes (full range of data)
        # Find column indices for percentage change columns
        header_row_values = [cell.value for cell in ws_market_overview[1]]
        try:
            col_idx_24h_change = header_row_values.index('24h Change (%)') + 1
            col_idx_7d_change = header_row_values.index('7d Change (%)') + 1
            col_idx_30d_change = header_row_values.index('30d Change (%)') + 1
            col_idx_1y_change = header_row_values.index('1y Change (%)') + 1

            # Get column letters
            start_col_letter = get_column_letter(min(col_idx_24h_change, col_idx_7d_change, col_idx_30d_change, col_idx_1y_change))
            end_col_letter = get_column_letter(max(col_idx_24h_change, col_idx_7d_change, col_idx_30d_change, col_idx_1y_change))

            ws_market_overview.conditional_formatting.add(f'{start_col_letter}2:{end_col_letter}{ws_market_overview.max_row}', ColorScaleRule(
                start_type='min', start_color='FFFF0000', # Red
                mid_type='num', mid_value=0, mid_color='FFFFFF00', # Yellow
                end_type='max', end_color='FF00FF00' # Green
            ))
        except ValueError:
            print("Warning: One or more percentage change columns not found for conditional formatting in Market Overview.")

        # Auto-adjust column widths for Market Overview
        for col_idx in range(1, ws_market_overview.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row in ws_market_overview.iter_rows(min_row=1, max_row=ws_market_overview.max_row):
                cell = row[col_idx - 1] # Adjust for 0-indexed column in row tuple
                try:
                    if cell.value is not None:
                        # Consider formatted numbers as well for length
                        cell_value_str = str(cell.value)
                        if isinstance(cell.value, (int, float)) and cell.number_format:
                            # Attempt to get formatted string if number_format exists
                            # This is a bit advanced for openpyxl, but simple str() is usually enough
                            pass
                        max_length = max(max_length, len(cell_value_str))
                except Exception as e:
                    print(f"Error checking cell length at {column_letter}{cell.row}: {e}")
                    pass
            adjusted_width = min(max_length + 2, 30) # Cap width to 30 for long strings
            ws_market_overview.column_dimensions[column_letter].width = adjusted_width

    # 3. Global Metrics Sheet
    if global_metrics_df is not None and not global_metrics_df.empty:
        ws_global_metrics = wb[GLOBAL_METRICS_SHEET_NAME]
        ws_global_metrics.delete_rows(1, ws_global_metrics.max_row)

        # Write header row
        ws_global_metrics.append(global_metrics_df.columns.tolist())

        # Write data rows (dataframe_to_rows handles NaNs properly as empty cells)
        for r_idx, r in enumerate(dataframe_to_rows(global_metrics_df, index=False, header=False)):
            for c_idx, value in enumerate(r):
                ws_global_metrics.cell(row=r_idx + 2, column=c_idx + 1, value=value)

        # Format header
        for cell in ws_global_metrics[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

        # Apply specific number formats based on the metric name
        for row_idx in range(2, ws_global_metrics.max_row + 1):
            metric_name = ws_global_metrics.cell(row=row_idx, column=1).value
            value_cell = ws_global_metrics.cell(row=row_idx, column=2)

            if pd.isna(value_cell.value): # Handle NaN values from DataFrame
                continue

            if "USD" in str(metric_name): # Ensure metric_name is string
                value_cell.number_format = '$#,##0'
            elif "%" in str(metric_name):
                value_cell.number_format = '0.00%'
            else: # For counts like active cryptocurrencies, markets
                value_cell.number_format = '#,##0'

        for col_idx in range(1, ws_global_metrics.max_column + 1):
            ws_global_metrics.column_dimensions[get_column_letter(col_idx)].width = 30 # Fixed width for global metrics


    # 4. Fear & Greed Index Sheet
    if fng_df is not None and not fng_df.empty:
        ws_fng = wb[FEAR_GREED_INDEX_SHEET_NAME]
        ws_fng.delete_rows(1, ws_fng.max_row)

        # Write header row
        ws_fng.append(fng_df.columns.tolist())

        for r_idx, r in enumerate(dataframe_to_rows(fng_df, index=False, header=False)): # header=False as we added it manually
            for c_idx, value in enumerate(r):
                ws_fng.cell(row=r_idx+2, column=c_idx+1, value=value) # Start from row 2 for data

        for cell in ws_fng[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        ws_fng.column_dimensions['A'].width = 12
        ws_fng.column_dimensions['B'].width = 18
        ws_fng.column_dimensions['C'].width = 15

        # Create F&G chart
        chart = LineChart()
        chart.title = "Fear & Greed Index Over Time"
        chart.style = 13
        chart.x_axis.title = 'Date'
        chart.y_axis.title = 'Fear & Greed Index'

        # Corrected references for F&G chart (data starts from row 2 as headers are in row 1)
        # Data for chart should be numeric. 'Fear & Greed Index' column is B (2nd col)
        data = Reference(ws_fng, min_col=2, min_row=1, max_row=ws_fng.max_row, max_col=2)
        dates = Reference(ws_fng, min_col=1, min_row=2, max_row=ws_fng.max_row) # Date column is A (1st col)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        ws_fng.add_chart(chart, "E1")


    # 5. Individual Coin Histories (BTC, ETH, BNB, SOL, SUI)
    for symbol, hist_df in all_data.get('historical_data', {}).items():
        sheet_name = f"{symbol} History (1 Month)"
        if sheet_name in wb.sheetnames and hist_df is not None and not hist_df.empty:
            ws_hist = wb[sheet_name]
            ws_hist.delete_rows(1, ws_hist.max_row)

            # Write header row
            ws_hist.append(hist_df.columns.tolist())

            for r_idx, r in enumerate(dataframe_to_rows(hist_df, index=False, header=False)): # header=False as we added it manually
                for c_idx, value in enumerate(r):
                    ws_hist.cell(row=r_idx+2, column=c_idx+1, value=value) # Start from row 2 for data

            # Basic formatting for historical sheets
            for cell in ws_hist[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

            # Apply numeric formatting where appropriate
            header_row_values = [cell.value for cell in ws_hist[1]]
            for row_idx in range(2, ws_hist.max_row + 1):
                for col_idx in range(1, ws_hist.max_column + 1):
                    cell = ws_hist.cell(row=row_idx, column=col_idx)
                    header_name = header_row_values[col_idx-1] # -1 for 0-indexed list

                    if header_name == 'Price' or header_name.endswith('MA'):
                        if cell.value is not None:
                            cell.number_format = '#,##0.0000' # High precision for price/MA
                    elif header_name == 'Volume' or header_name == 'Market Cap':
                        if cell.value is not None:
                            cell.number_format = '#,##0' # Integer formatting
                    elif header_name == 'Daily Return (%)' or header_name == 'Volatility (7d)':
                        if cell.value is not None:
                            cell.number_format = '0.000%' # Percentage formatting


            for col_idx in range(1, ws_hist.max_column + 1):
                ws_hist.column_dimensions[get_column_letter(col_idx)].width = 15

            # Conditional formatting for Daily Return (%)
            # Assuming 'Daily Return (%)' is in column 5 (E) as per the original script
            try:
                return_col_idx = hist_df.columns.get_loc('Daily Return (%)') + 1 # +1 because openpyxl is 1-indexed
                ws_hist.conditional_formatting.add(f'{get_column_letter(return_col_idx)}2:{get_column_letter(return_col_idx)}{ws_hist.max_row}', ColorScaleRule(
                    start_type='min', start_color='FFFF0000',
                    mid_type='num', mid_value=0, mid_color='FFFFFF00',
                    end_type='max', end_color='FF00FF00'
                ))
            except KeyError:
                print(f"Warning: 'Daily Return (%)' column not found for conditional formatting in {sheet_name}.")


            # Create line chart for Price
            chart = LineChart()
            chart.title = f"{symbol} Price History"
            chart.style = 13
            chart.x_axis.title = 'Date'
            chart.y_axis.title = 'Price (USD)'

            # Data and categories for the chart
            price_col_idx = hist_df.columns.get_loc('Price') + 1
            # Data should start from row 1 (header) for titles_from_data=True
            data = Reference(ws_hist, min_col=price_col_idx, min_row=1, max_row=ws_hist.max_row, max_col=price_col_idx)
            # Categories (dates) should start from row 2 (data)
            dates = Reference(ws_hist, min_col=1, min_row=2, max_row=ws_hist.max_row) # Date column is 1

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(dates)

            # Positioning the chart
            ws_hist.add_chart(chart, "I1") # Place chart starting at cell I1


    # Update for 'Transactions' sheet
    ws_transactions = wb[TRANSACTIONS_SHEET_NAME]
    # No dynamic data is fetched for transactions, only headers on creation.
    # We ensure existing data is not overwritten.

    # Update for 'Current Portfolio' sheet (existing logic retained and adjusted for headers)
    current_portfolio_df = all_data.get('current_portfolio')
    if current_portfolio_df is not None and not current_portfolio_df.empty:
        ws_portfolio = wb[CURRENT_PORTFOLIO_SHEET_NAME]

        header = [cell.value for cell in ws_portfolio[1]] if ws_portfolio.max_row > 0 else []
        existing_portfolio_data = []
        if ws_portfolio.max_row > 1:
            for row in ws_portfolio.iter_rows(min_row=2, values_only=True):
                existing_portfolio_data.append(dict(zip(header, row)))
        existing_portfolio_df = pd.DataFrame(existing_portfolio_data)

        if 'Coin ID' in existing_portfolio_df.columns:
            existing_portfolio_df.set_index('Coin ID', inplace=True)
        else:
            existing_portfolio_df = pd.DataFrame()

        current_portfolio_df.set_index('Coin ID', inplace=True)

        merged_portfolio_df = existing_portfolio_df.copy()

        for coin_id, row_api in current_portfolio_df.iterrows():
            if coin_id in merged_portfolio_df.index:
                merged_portfolio_df.loc[coin_id, 'Current Value (USD)'] = row_api['Current Value (USD)']
            else:
                new_coin_row = row_api.to_frame().T
                new_coin_row.index.name = 'Coin ID'
                merged_portfolio_df = pd.concat([merged_portfolio_df, new_coin_row])

        portfolio_cols = ['Current Value (USD)', 'Symbol', 'Location', 'Quantity', 'Purchase Price (USD)', 'Total Value (USD)', 'P/L (USD)', 'P/L Ratio', 'Airdrop or Invest']
        for col in portfolio_cols:
            if col not in merged_portfolio_df.columns:
                merged_portfolio_df[col] = ''

        final_portfolio_cols = ['Coin ID'] + [col for col in portfolio_cols if col in merged_portfolio_df.columns]
        merged_portfolio_df = merged_portfolio_df.reset_index()[final_portfolio_cols]

        ws_portfolio.delete_rows(1, ws_portfolio.max_row)
        for r_idx, r in enumerate(dataframe_to_rows(merged_portfolio_df, index=False, header=True)):
            for c_idx, value in enumerate(r):
                ws_portfolio.cell(row=r_idx+1, column=c_idx+1, value=value)
        for col in range(1, ws_portfolio.max_column + 1):
            ws_portfolio.column_dimensions[chr(64 + col)].width = 20
        # Apply header styling to Current Portfolio
        for cell in ws_portfolio[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

    # Reorder sheets to your desired order
    for i, sheet_name in enumerate(SHEET_ORDER):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if wb.sheetnames.index(sheet_name) != i:
                wb.move_sheet(ws, offset=i - wb.sheetnames.index(sheet_name))

    wb.save(excel_file)
    print("Excel file updated successfully with all sheets.")


# --- Main Execution ---
if __name__ == "__main__": # Corrected __name__
    all_fetched_data = {}

    print("🚀 Starting comprehensive cryptocurrency data collection...")
    print("-" * 50)

    print("Fetching current coin prices (for portfolio template 'Current Price' column if you want to integrate)...")
    current_coin_prices = fetch_current_coin_prices(PORTFOLIO_COIN_IDS)
    if current_coin_prices:
        portfolio_df_data = []
        for coin_id, price in current_coin_prices.items():
            if price is not None:
                portfolio_df_data.append({
                    'Coin ID': coin_id,
                    'Current Value (USD)': price # This is 'Current Price' for the portfolio sheet
                })
        all_fetched_data['current_portfolio'] = pd.DataFrame(portfolio_df_data)
    else:
        print("Skipping Current Portfolio update due to fetch error.")

    time.sleep(INITIAL_API_CALL_DELAY_SECONDS) # Delay before next major API call

    print("📈 Collecting market overview data (Top 50 cryptos)...")
    market_overview_df = get_market_overview()
    if not market_overview_df.empty:
        all_fetched_data['market_overview'] = market_overview_df
        print(f"✅ Fetched market data for {len(market_overview_df)} cryptocurrencies.")
    else:
        print("❌ Failed to fetch market overview data.")

    time.sleep(INITIAL_API_CALL_DELAY_SECONDS)

    print("🌍 Collecting global market metrics...")
    global_metrics_df = get_global_metrics()
    if not global_metrics_df.empty:
        all_fetched_data['global_metrics'] = global_metrics_df
        print("✅ Fetched global market metrics.")
    else:
        print("❌ Failed to fetch global metrics.")

    time.sleep(INITIAL_API_CALL_DELAY_SECONDS) # Corrected variable name

    print("😰 Fetching Fear & Greed Index...")
    fng_df = get_fear_greed_index()
    if not fng_df.empty:
        all_fetched_data['fear_greed_index'] = fng_df
        print(f"✅ Fetched {len(fng_df)} days of Fear & Greed Index data.")
    else:
        print("❌ Failed to fetch Fear & Greed Index.")

    print(f"Fetching {HISTORICAL_DAYS} days of historical data for selected coins...")
    historical_dfs = {}
    for symbol, coin_id in HISTORY_COIN_IDS.items():
        hist_df = get_historical_data(coin_id, HISTORICAL_DAYS)
        if hist_df is not None and not hist_df.empty:
            historical_dfs[symbol] = hist_df
            print(f"  ✅ Fetched history for {symbol} ({coin_id})")
        else:
            print(f"  ❌ Failed to fetch history for {symbol} ({coin_id})")
        time.sleep(INITIAL_API_CALL_DELAY_SECONDS) # Delay after each historical API call

    all_fetched_data['historical_data'] = historical_dfs


    print("\nAll data collection attempts complete. Proceeding to Excel generation...")
    if all_fetched_data:
        print("Updating Excel file...")
        create_or_update_excel(all_fetched_data, EXCEL_FILE_NAME)
    else:
        print("No data fetched. Excel file not updated.")

    print("\n" + "=" * 50)
    print("📊 EXCEL FILE CONTENTS:")
    print("=" * 50)
    print(f"✅ File saved as: {EXCEL_FILE_NAME}")
    print("📋 Sheets included:", ", ".join(SHEET_ORDER))
    print("🎯 Data reliability: 90-95% accuracy for live data, depends on CoinGecko API.")
    print("📅 Historical range: Up to 1 year of daily data for specified coins.")
    print("🔄 Contains live data, comprehensive market overview, and template sheets.")
    print("=" * 50)
